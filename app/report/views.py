from rest_framework import generics
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response

from app.sales.models import Sale
from app.sales.filters import SaleFilter

from app.inventory.models import Item

from .serializers import ReportSaleSerializer

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import shutil
from django.conf import settings
from django.utils.timezone import now


class ReportSaleList(generics.ListAPIView):
    """
    get:
        Url con reporte de ventas.
    """

    queryset = Sale.objects.all().order_by('date')
    serializer_class = ReportSaleSerializer
    filter_backends = (DjangoFilterBackend,)
    filter_class = SaleFilter

    def list(self, request):
        queryset = self.filter_queryset(self.get_queryset())

        path = self.createSheet(queryset)
        filename = 'RESUMEN_' \
            + now().strftime("%Y%m%d%H%M%S") \
            + '.xlsx'
        new_path = settings.MEDIA_ROOT \
            + filename
        shutil.move(path, new_path)

        url = settings.MEDIA_URL + filename
        data = {
            'url': url
        }

        serializer = self.get_serializer(data)
        return Response(serializer.data)

    def createSheet(self, queryset):
        dir = '/tmp/'
        filename = 'resumen.xlsx'
        path = dir + filename

        items = Item.objects.all()
        items_sold = {}
        for item in items:
            items_sold[item.id] = 0

        total = 0
        active_sales = 0

        wb = Workbook()

        ws = wb.active
        ws.title = 'Resumen'
        wb.create_sheet(title='Ventas')
        wb.create_sheet(title='Productos vendidos')

        ws = wb['Ventas']

        row = 1
        col = 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 10

        col = 1
        ws.cell(row, col, value='Fecha').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Total').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Número de artículos').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Medio de pago').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Nula').style = 'Accent1'

        row += 1
        for sale in queryset:

            col = 1
            ws.cell(row, col, value=sale.date)

            col += 1
            ws.cell(row, col, value=sale.total)
            if sale.active:
                total += sale.total

            col += 1
            n = 0
            for saleitem in sale.saleitem_set.all():
                n += saleitem.quantity
                if sale.active and saleitem.item:
                    items_sold[saleitem.item.id] += saleitem.quantity
            ws.cell(row, col, value=n)

            col += 1
            if sale.paymentmethod:
                ws.cell(row, col, value=sale.paymentmethod.name)

            col += 1
            if sale.active:
                ws.cell(row, col, value='Vigente')
                active_sales += 1
            else:
                ws.cell(row, col, value='Nula').style = 'Bad'
            row += 1

        col = 2
        ws.cell(row, col, value='TOTAL').style = 'Accent1'

        col += 1
        ws.cell(row, col, value=total).style = 'Accent1'

        ws = wb['Resumen']

        row = 1
        col = 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 10

        row = 1
        col = 1
        ws.cell(row, col, value='Resumen').style = 'Accent1'

        row += 1
        col = 1
        ws.cell(row, col, value='Número de ventas válidas')

        col += 1
        ws.cell(row, col, value=active_sales)

        row += 1
        col = 1
        ws.cell(row, col, value='Número de ventas anuladas')

        col += 1
        ws.cell(row, col, value=queryset.count() - active_sales)

        row += 1
        col = 1
        ws.cell(row, col, value='Total')

        col += 1
        ws.cell(row, col, value=total)

        ws = wb['Productos vendidos']

        row = 1
        col = 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 20
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 20

        col = 1
        ws.cell(row, col, value='Código').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Nombre').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Cantidad Vendida').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='En Stock').style = 'Accent1'

        col += 1
        ws.cell(row, col, value='Stock mínimo').style = 'Accent1'

        row += 1
        for item in items:

            col = 1
            ws.cell(row, col, value=item.bar_code)

            col += 1
            ws.cell(row, col, value=item.name)

            col += 1
            ws.cell(row, col, value=items_sold[item.id])

            col += 1
            if item.stock < item.stock_min:
                ws.cell(row, col, value=item.stock).style = 'Bad'
            else:
                ws.cell(row, col, value=item.stock)

            col += 1
            ws.cell(row, col, value=item.stock_min)
            row += 1

        wb.save(path)
        return path
